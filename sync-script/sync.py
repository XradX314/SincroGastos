"""
SincroGastos — Sincronización bidireccional Excel ↔ Supabase
=============================================================
100% tkinter: funciona como .exe sin consola ni Python instalado.

Empaquetar:
    pip install -r requirements.txt
    python build.py          ← genera dist/SincroGastos-Sync.exe
"""

from __future__ import annotations

import base64
import hashlib
import json
import logging
import os
import shutil
import sys
import tkinter as tk
from copy import copy
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk
from typing import Any, TypedDict

import openpyxl
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes as crypto_hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from supabase import Client, create_client

# ─────────────────────────────────────────────────────────────
# Ruta del config — junto al .exe en producción, junto al script en dev
# ─────────────────────────────────────────────────────────────

def _app_dir() -> Path:
    """Directorio del ejecutable (frozen) o del script (dev)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _resource_path(rel: str) -> Path:
    """Path a un recurso empaquetado (dentro de _MEIPASS si es .exe, junto al script en dev)."""
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS) / rel  # type: ignore[attr-defined]
    return Path(__file__).parent / rel


CONFIG_FILE = _app_dir() / "config.json"

# ─────────────────────────────────────────────────────────────
# Tipos
# ─────────────────────────────────────────────────────────────

class ExpenseRow(TypedDict):
    categoria: str
    subcategoria: str
    fecha: str
    detalle: str
    importe: float
    origen: str
    hash_unico: str


class Config(TypedDict):
    supabase_url: str
    supabase_key: str
    excel_path: str
    encrypted_token: str
    salt: str


# ─────────────────────────────────────────────────────────────
# Logging (a archivo junto al .exe para poder ver errores)
# ─────────────────────────────────────────────────────────────

log_path = _app_dir() / "sincrogastos.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("sincrogastos")

# Categorías aceptadas (sin restricción: cualquier texto no vacío es válido)
CATEGORIA_VALIDAS: set[str] = set()  # vacío = aceptar todo

MONTH_TABS: dict[str, int] = {
    "01 Enero": 1, "02 Febrero": 2, "03 Marzo": 3,
    "04 Abril": 4, "05 Mayo": 5, "06 Junio": 6,
    "07 Julio": 7, "08 Agosto": 8, "09 Septiembre": 9,
    "10 Octubre": 10, "11 Noviembre": 11, "12 Diciembre": 12,
}


# ─────────────────────────────────────────────────────────────
# JWT helpers
# ─────────────────────────────────────────────────────────────

def _user_id_from_token(token: str) -> str:
    """Extrae el user_id (sub) del JWT sin hacer ninguna llamada de red."""
    payload_b64 = token.split('.')[1]
    padding = 4 - len(payload_b64) % 4
    if padding != 4:
        payload_b64 += '=' * padding
    payload = json.loads(base64.urlsafe_b64decode(payload_b64))
    return payload['sub']


# ─────────────────────────────────────────────────────────────
# Cifrado
# ─────────────────────────────────────────────────────────────

def _derive_key(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(
        algorithm=crypto_hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=390_000,
    )
    return base64.urlsafe_b64encode(kdf.derive(password.encode()))


def encrypt_token(token: str, password: str) -> tuple[str, str]:
    salt = os.urandom(16)
    key = _derive_key(password, salt)
    encrypted = Fernet(key).encrypt(token.encode())
    return base64.urlsafe_b64encode(encrypted).decode(), salt.hex()


def decrypt_token(encrypted_b64: str, salt_hex: str, password: str) -> str:
    salt = bytes.fromhex(salt_hex)
    key = _derive_key(password, salt)
    encrypted = base64.urlsafe_b64decode(encrypted_b64.encode())
    return Fernet(key).decrypt(encrypted).decode()


# ─────────────────────────────────────────────────────────────
# Hash (mismo algoritmo que el frontend)
# ─────────────────────────────────────────────────────────────

def generate_hash(row: ExpenseRow) -> str:
    raw = "|".join([
        row["categoria"],
        row["subcategoria"],
        row["fecha"],
        row.get("detalle") or "",
        str(row["importe"]),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ─────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────

def load_config() -> Config | None:
    if not CONFIG_FILE.exists():
        return None
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)  # type: ignore[return-value]


def save_config(cfg: Config) -> None:
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ─────────────────────────────────────────────────────────────
# Wizard tkinter (sin consola — funciona en .exe windowed)
# ─────────────────────────────────────────────────────────────

class WizardDialog(tk.Toplevel):
    """Diálogo de primera configuración completamente en tkinter."""

    def __init__(self, parent: tk.Tk) -> None:
        super().__init__(parent)
        self.title("SincroGastos — Primera configuración")
        self.resizable(False, False)
        self.grab_set()
        self.result: Config | None = None
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build_ui(self) -> None:
        BG = "#1E1E2E"
        FG = "#CDD6F4"
        ENTRY_BG = "#313244"
        self.configure(bg=BG)

        tk.Label(self, text="Configuración inicial", font=("Inter", 14, "bold"),
                 bg=BG, fg=FG).pack(pady=(16, 4))
        tk.Label(self, text="Completá estos datos una sola vez.", font=("Inter", 9),
                 bg=BG, fg="#89B4FA").pack(pady=(0, 12))

        form = tk.Frame(self, bg=BG)
        form.pack(padx=24, fill="x")

        fields = [
            ("URL del proyecto Supabase", "https://xxxx.supabase.co", False),
            ("Anon Key de Supabase", "", False),
            ("Email de tu cuenta SincroGastos", "", False),
            ("Contraseña de SincroGastos", "", True),
        ]
        self._entries: list[tk.Entry] = []
        for label, placeholder, secret in fields:
            tk.Label(form, text=label, font=("Inter", 9), bg=BG, fg=FG,
                     anchor="w").pack(fill="x", pady=(6, 1))
            e = tk.Entry(form, font=("Inter", 10), bg=ENTRY_BG, fg=FG,
                         insertbackground=FG, relief="flat",
                         show="•" if secret else "")
            if placeholder:
                e.insert(0, placeholder)
                e.bind("<FocusIn>", lambda ev, p=placeholder, en=e: (
                    en.delete(0, "end") if en.get() == p else None
                ))
            e.pack(fill="x", ipady=6)
            self._entries.append(e)

        # Selector de Excel
        tk.Label(form, text="Archivo gastos.xlsx", font=("Inter", 9),
                 bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(10, 1))
        excel_row = tk.Frame(form, bg=BG)
        excel_row.pack(fill="x")
        self._excel_var = tk.StringVar()
        tk.Entry(excel_row, textvariable=self._excel_var, font=("Inter", 9),
                 bg=ENTRY_BG, fg=FG, insertbackground=FG, relief="flat",
                 state="readonly").pack(side="left", fill="x", expand=True, ipady=6)
        tk.Button(excel_row, text="Buscar…", command=self._pick_excel,
                  bg="#4F6AF5", fg="white", relief="flat", padx=8,
                  cursor="hand2").pack(side="left", padx=(6, 0))

        # Botones
        btn_row = tk.Frame(self, bg=BG)
        btn_row.pack(pady=16, padx=24, fill="x")
        tk.Button(btn_row, text="Cancelar", command=self._on_cancel,
                  bg="#313244", fg=FG, relief="flat", padx=12, pady=6,
                  cursor="hand2").pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text="Guardar y conectar", command=self._on_save,
                  bg="#4F6AF5", fg="white", font=("Inter", 10, "bold"),
                  relief="flat", padx=12, pady=6, cursor="hand2").pack(side="right")

    def _pick_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleccioná el archivo gastos.xlsx",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
        )
        if path:
            self._excel_var.set(path)

    def _on_save(self) -> None:
        url, key, email, password = (e.get().strip() for e in self._entries)
        excel = self._excel_var.get().strip()

        if not all([url, key, email, password, excel]):
            messagebox.showwarning("Faltan datos", "Completá todos los campos.", parent=self)
            return

        # Autenticar para obtener token
        try:
            client: Client = create_client(url, key)
            response = client.auth.sign_in_with_password({"email": email, "password": password})
            if not response.session:
                raise ValueError("Sin sesión")
        except Exception as exc:
            messagebox.showerror("Error de autenticación",
                                 f"No se pudo conectar con Supabase:\n{exc}", parent=self)
            return

        enc_token, salt = encrypt_token(response.session.access_token, password)
        self.result = Config(
            supabase_url=url,
            supabase_key=key,
            excel_path=excel,
            encrypted_token=enc_token,
            salt=salt,
        )
        save_config(self.result)
        self.destroy()

    def _on_cancel(self) -> None:
        self.result = None
        self.destroy()


def run_wizard(parent: tk.Tk) -> Config | None:
    dlg = WizardDialog(parent)
    parent.wait_window(dlg)
    return dlg.result


# ─────────────────────────────────────────────────────────────
# Lectura del Excel
# ─────────────────────────────────────────────────────────────

def _parse_date(cell_value: Any) -> str | None:
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    if hasattr(cell_value, "strftime"):
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, (int, float)):
        # Número de serie de Excel → fecha
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(cell_value).strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(cell_value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(cell_value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _fallback_date(sheet_name: str) -> str:
    """Usa el 1° del mes de la pestaña como fecha si la celda está vacía."""
    month = MONTH_TABS.get(sheet_name, 1)
    year = datetime.now().year
    return f"{year}-{month:02d}-01"


def _detect_col_offset(all_rows: list[tuple[Any, ...]]) -> int:
    """Detecta en qué columna empieza CATEGORÍA (0=A, 1=B, etc.)."""
    keywords = {"CATEGORÍA", "CATEGORIA", "CATEGORY"}
    for row in all_rows:
        for col_idx, cell in enumerate(row):
            if cell and str(cell).strip().upper().replace("�", "Í") in keywords:
                return col_idx
    return 1  # default: columna B (layout SincroGastos)


def _is_total_row(cells: list[Any], col: int) -> bool:
    cat = cells[col] if len(cells) > col else None
    if cat is None:
        return True
    s = str(cat).strip()
    return s == "" or s.upper().startswith("TOTAL") or s == "x"


def read_excel(excel_path: str) -> list[ExpenseRow]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    rows: list[ExpenseRow] = []

    for sheet_name in MONTH_TABS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        col = _detect_col_offset(all_rows)
        header_idx = 0
        for i, row in enumerate(all_rows):
            if len(row) > col and row[col] and "CATEGOR" in str(row[col]).upper():
                header_idx = i + 1
                break

        for raw_row in all_rows[header_idx:]:
            cells = list(raw_row)
            if len(cells) < col + 5:
                continue
            if _is_total_row(cells, col):
                continue

            categoria    = str(cells[col]).strip()   if cells[col]   else ""
            subcategoria = str(cells[col+1]).strip() if cells[col+1] else ""
            fecha_raw    = cells[col+2]
            detalle      = str(cells[col+3]).strip() if cells[col+3] else ""
            importe_raw  = cells[col+4]

            if not categoria or not subcategoria:
                continue

            fecha = _parse_date(fecha_raw)
            if not fecha:
                fecha = _fallback_date(sheet_name)
                logger.warning("Fecha vacía en '%s' → usando %s", sheet_name, fecha)

            try:
                importe = float(str(importe_raw).replace(",", ".").replace("$", "").strip())
            except (ValueError, TypeError):
                logger.warning("Importe inválido: %r", importe_raw)
                continue

            expense: ExpenseRow = {
                "categoria": categoria,
                "subcategoria": subcategoria,
                "fecha": fecha,
                "detalle": detalle,
                "importe": importe,
                "origen": "excel",
                "hash_unico": "",
            }
            expense["hash_unico"] = generate_hash(expense)
            rows.append(expense)

    wb.close()
    logger.info("Excel leído: %d filas.", len(rows))
    return rows


# ─────────────────────────────────────────────────────────────
# Supabase
# ─────────────────────────────────────────────────────────────

def get_supabase_client(cfg: Config, password: str) -> Client:
    client: Client = create_client(cfg["supabase_url"], cfg["supabase_key"])
    try:
        token = decrypt_token(cfg["encrypted_token"], cfg["salt"], password)
        client.auth.set_session(token, token)
        logger.info("Sesión restaurada.")
    except Exception as exc:
        logger.warning("Token expirado (%s). Reautenticando...", exc)
        raise RuntimeError("Token expirado. Usá el botón 'Reconfigurar'.") from exc
    return client


def fetch_cloud_expenses(client: Client) -> list[dict[str, Any]]:
    result = client.table("expenses").select("*").execute()
    data: list[dict[str, Any]] = result.data or []
    logger.info("Nube: %d gastos.", len(data))
    return data


def upload_expenses(client: Client, rows: list[ExpenseRow], user_id: str) -> int:
    if not rows:
        return 0
    uploaded = 0
    for i in range(0, len(rows), 100):
        batch = rows[i:i+100]
        result = client.table("expenses").upsert(
            [{**dict(r), "user_id": user_id} for r in batch],
            on_conflict="hash_unico",
        ).execute()
        uploaded += len(result.data or [])
    logger.info("Subidos: %d.", uploaded)
    return uploaded


# ─────────────────────────────────────────────────────────────
# Escritura en Excel
# ─────────────────────────────────────────────────────────────

def _month_from_date(date_str: str) -> int:
    return int(date_str[5:7])


def _tab_name_for_month(month: int) -> str | None:
    return next((n for n, m in MONTH_TABS.items() if m == month), None)


def _find_total_row(ws: Any) -> int:
    """
    Busca la fila del sentinel 'x' o de totales buscando en todas las columnas.
    Los nuevos gastos se insertan ANTES de esa fila.
    """
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, 10):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val).strip()
            if s.upper().startswith("TOTAL") or s == "x":
                return row_idx
    return ws.max_row + 1


def backup_excel(excel_path: str) -> str:
    src = Path(excel_path)
    dst = src.parent / f"{src.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{src.suffix}"
    shutil.copy2(src, dst)
    logger.info("Backup: %s", dst)
    return str(dst)


def write_expenses_to_excel(excel_path: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    backup_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    written = 0

    by_tab: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        month = _month_from_date(row["fecha"])
        tab = _tab_name_for_month(month)
        if tab and tab in wb.sheetnames:
            by_tab.setdefault(tab, []).append(row)

    for tab_name, tab_rows in by_tab.items():
        ws = wb[tab_name]
        total_idx = _find_total_row(ws)
        # Detectar columna de inicio (B=2 en este Excel)
        all_ws_rows = list(ws.iter_rows(values_only=True))
        col_start = _detect_col_offset(all_ws_rows) + 1  # convertir a 1-indexed

        for row_data in tab_rows:
            ws.insert_rows(total_idx)
            vals = [
                row_data.get("categoria", ""),
                row_data.get("subcategoria", ""),
                row_data.get("fecha", ""),
                row_data.get("detalle", ""),
                row_data.get("importe", 0),
            ]
            for offset, val in enumerate(vals):
                c = col_start + offset
                new_cell = ws.cell(row=total_idx, column=c, value=val)
                src_cell = ws.cell(row=total_idx - 1, column=c)
                if src_cell.has_style:
                    new_cell.font = copy(src_cell.font)
                    new_cell.border = copy(src_cell.border)
                    new_cell.fill = copy(src_cell.fill)
                    new_cell.number_format = src_cell.number_format
                    new_cell.alignment = copy(src_cell.alignment)
            total_idx += 1
            written += 1

    wb.save(excel_path)
    wb.close()
    logger.info("Excel guardado con %d filas nuevas.", written)
    return written


# ─────────────────────────────────────────────────────────────
# Algoritmo de merge
# ─────────────────────────────────────────────────────────────

class SyncResult:
    def __init__(self) -> None:
        self.uploaded = 0
        self.downloaded = 0
        self.conflicts = 0
        self.errors: list[str] = []


def sync(cfg: Config, client: Client, user_id: str, log_cb: Any = None) -> SyncResult:
    result = SyncResult()

    def log(msg: str) -> None:
        logger.info(msg)
        if log_cb:
            log_cb(msg)

    log("📖 Leyendo Excel...")
    try:
        excel_rows = read_excel(cfg["excel_path"])
    except Exception as exc:
        result.errors.append(str(exc))
        log(f"❌ {exc}")
        return result

    set_excel = {r["hash_unico"]: r for r in excel_rows}

    log("☁️  Consultando Supabase...")
    try:
        cloud_rows = fetch_cloud_expenses(client)
    except Exception as exc:
        result.errors.append(str(exc))
        log(f"❌ {exc}")
        return result

    set_cloud = {r["hash_unico"]: r for r in cloud_rows}

    only_in_cloud = [r for h, r in set_cloud.items() if h not in set_excel]
    only_in_excel = [r for h, r in set_excel.items() if h not in set_cloud]  # type: ignore[misc]

    log(f"🔍 Solo en nube: {len(only_in_cloud)} | Solo en Excel: {len(only_in_excel)}")

    if only_in_excel:
        log(f"⬆️  Subiendo {len(only_in_excel)} gastos...")
        try:
            result.uploaded = upload_expenses(client, list(only_in_excel), user_id)  # type: ignore[arg-type]
        except Exception as exc:
            result.errors.append(str(exc))
            log(f"❌ {exc}")

    if only_in_cloud:
        log(f"⬇️  Descargando {len(only_in_cloud)} gastos al Excel...")
        try:
            result.downloaded = write_expenses_to_excel(cfg["excel_path"], only_in_cloud)
        except Exception as exc:
            result.errors.append(str(exc))
            log(f"❌ {exc}")

    log("✅ Sincronización completada.")
    return result


# ─────────────────────────────────────────────────────────────
# GUI principal
# ─────────────────────────────────────────────────────────────

class SyncApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.cfg: Config | None = load_config()
        self.client: Client | None = None
        self.title("SincroGastos")
        self.resizable(False, False)
        self.geometry("600x520")
        self._setup_ui()
        self._apply_theme()

        # Si no hay config, abrir wizard al arrancar
        if not self.cfg:
            self.after(200, self._open_wizard)
        else:
            self._refresh_status()

    # ── Tema ─────────────────────────────────────────────────

    def _apply_theme(self) -> None:
        self.BG    = "#1E1E2E"
        self.FG    = "#CDD6F4"
        self.BLUE  = "#4F6AF5"
        self.configure(bg=self.BG)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TProgressbar", troughcolor=self.BG, background=self.BLUE)

    # ── UI ───────────────────────────────────────────────────

    def _setup_ui(self) -> None:
        pad = {"padx": 16, "pady": 6}

        # Cabecera con logo
        try:
            from PIL import Image, ImageTk
            logo_path = _resource_path("logo.png")
            img = Image.open(logo_path)
            target_h = 52
            ratio = target_h / img.height
            img = img.resize((int(img.width * ratio), target_h), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(self, image=self._logo_img, bg="#1E1E2E").pack(
                padx=16, pady=(12, 0), anchor="w"
            )
        except Exception:
            tk.Label(self, text="SincroGastos", font=("Inter", 17, "bold"),
                     fg="#CDD6F4", bg="#1E1E2E").pack(padx=16, pady=(16, 2), anchor="w")

        self.status_label = tk.Label(
            self, text="Sin configuración", font=("Inter", 9),
            fg="#89B4FA", bg="#1E1E2E", anchor="w",
        )
        self.status_label.pack(padx=16, anchor="w")

        ttk.Separator(self).pack(fill="x", pady=8)

        # Contraseña
        pw_frame = tk.Frame(self, bg="#1E1E2E")
        pw_frame.pack(fill="x", **pad)
        tk.Label(pw_frame, text="Contraseña:", font=("Inter", 10),
                 fg="#CDD6F4", bg="#1E1E2E").pack(side="left")
        self.pw_var = tk.StringVar()
        tk.Entry(pw_frame, textvariable=self.pw_var, show="•", width=28,
                 font=("Inter", 10), bg="#313244", fg="#CDD6F4",
                 insertbackground="#CDD6F4", relief="flat").pack(side="left", padx=8, ipady=4)

        # Botones de acción
        btn_frame = tk.Frame(self, bg="#1E1E2E")
        btn_frame.pack(fill="x", padx=16, pady=4)

        self.btn_sync = tk.Button(
            btn_frame, text="▶  Sincronizar", command=self._start_sync,
            font=("Inter", 11, "bold"), bg="#4F6AF5", fg="white",
            activebackground="#3D57E0", relief="flat",
            padx=16, pady=8, cursor="hand2",
        )
        self.btn_sync.pack(side="left")

        tk.Button(
            btn_frame, text="⚙  Reconfigurar", command=self._open_wizard,
            font=("Inter", 9), bg="#313244", fg="#CDD6F4",
            activebackground="#45475A", relief="flat",
            padx=10, pady=8, cursor="hand2",
        ).pack(side="left", padx=(8, 0))

        # Barra de progreso
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=568)
        self.progress.pack(padx=16, pady=6)

        # Log
        self.log_text = scrolledtext.ScrolledText(
            self, height=14, font=("Cascadia Code", 9),
            bg="#181825", fg="#A6E3A1", insertbackground="#CDD6F4",
            relief="flat", wrap="word", state="disabled",
        )
        self.log_text.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # Botón abrir Excel (oculto hasta sincronizar)
        self.btn_open = tk.Button(
            self, text="📂  Abrir Excel", command=self._open_excel,
            font=("Inter", 10, "bold"), bg="#313244", fg="#CDD6F4",
            activebackground="#45475A", relief="flat",
            padx=12, pady=6, cursor="hand2",
        )

    def _refresh_status(self) -> None:
        if self.cfg:
            excel_name = Path(self.cfg["excel_path"]).name
            self.status_label.config(
                text=f"📄 {excel_name}  ·  🔗 {self.cfg['supabase_url']}"
            )
        else:
            self.status_label.config(text="Sin configuración — usá ⚙ Reconfigurar")

    # ── Wizard ───────────────────────────────────────────────

    def _open_wizard(self) -> None:
        cfg = run_wizard(self)
        if cfg:
            self.cfg = cfg
            self._refresh_status()
            self._log("✅ Configuración guardada.")

    # ── Sincronización ───────────────────────────────────────

    def _log(self, msg: str) -> None:
        self.log_text.configure(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {msg}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _start_sync(self) -> None:
        if not self.cfg:
            messagebox.showwarning("Sin configuración", "Usá ⚙ Reconfigurar primero.")
            return
        password = self.pw_var.get()
        if not password:
            messagebox.showwarning("Contraseña", "Ingresá tu contraseña.")
            return
        self.btn_sync.configure(state="disabled")
        self.progress.start(10)
        self._log("🔐 Autenticando...")
        self.after(50, lambda: self._do_sync(password))

    def _do_sync(self, password: str) -> None:
        try:
            self.client = get_supabase_client(self.cfg, password)  # type: ignore[arg-type]
            # Extraer user_id del JWT sin llamada de red adicional
            token = decrypt_token(self.cfg["encrypted_token"], self.cfg["salt"], password)  # type: ignore[arg-type]
            user_id = _user_id_from_token(token)
            self._log("✅ Autenticación exitosa.")
        except Exception as exc:
            self._log(f"❌ {exc}")
            self.progress.stop()
            self.btn_sync.configure(state="normal")
            messagebox.showerror("Error de autenticación", str(exc))
            return

        self._log("🔄 Sincronizando...")
        try:
            result = sync(self.cfg, self.client, user_id=user_id, log_cb=self._log)  # type: ignore[arg-type]
        except Exception as exc:
            self._log(f"❌ Error inesperado: {exc}")
            self.progress.stop()
            self.btn_sync.configure(state="normal")
            return

        self.progress.stop()
        self.btn_sync.configure(state="normal")

        self._log("─" * 48)
        self._log(f"✅ {result.uploaded} gastos subidos a la nube")
        self._log(f"✅ {result.downloaded} gastos descargados al Excel")
        if result.errors:
            for e in result.errors:
                self._log(f"❌ {e}")

        self.btn_open.pack(padx=16, pady=(0, 10))

    def _open_excel(self) -> None:
        if self.cfg:
            os.startfile(self.cfg["excel_path"])  # type: ignore[attr-defined]


# ─────────────────────────────────────────────────────────────
# Punto de entrada
# ─────────────────────────────────────────────────────────────

def main() -> None:
    app = SyncApp()
    app.mainloop()


if __name__ == "__main__":
    main()
